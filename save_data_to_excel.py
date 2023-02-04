import openpyxl
import gender_selection
from tkinter import messagebox
import os
import shutil
import pathlib


def write(registration_no, current_day_date, student, class_selection_combo, date_of_birth, religion, radio, skills,
          father_name, father_occupation, mother_name, mother_occupation, profile_image, b_certificate,
          form_four_certificate):
    try:
        reg_number = registration_no.get() 
        current_date = current_day_date.get()   
        student_name = student.get()        
        student_class = class_selection_combo.get()
        dob = date_of_birth.get()
        student_religion = religion.get()
        gender = gender_selection.gender_selected(radio)
        student_skills = skills.get()
        fathers_name = father_name.get()
        fathers_occupation = father_occupation.get()
        mothers_name = mother_name.get()
        mothers_occupation = mother_occupation.get()
        birth_certificate = b_certificate.get()
        f_certificate = form_four_certificate.get()
        
        if (reg_number == '' or current_date == '' or student_name == '' or dob == '' or student_religion == '' or
                student_skills == '' or student_class == '' or gender is None or father_occupation == ''
                or profile_image == '' or birth_certificate == '' or f_certificate == ''):
            messagebox.showerror(title='Error', message='Dont leave any field empty\nData was not saved!')

        else:
            excel_file_name = "./documents/Student_Registration_data.xlsx"
            excel_path = pathlib.Path(excel_file_name)
            if excel_path.exists():
                pass
            else:
                os.mkdir('./documents')

            excel_file = openpyxl.load_workbook(excel_file_name)
            sheet = excel_file.active
            
            reg_list = []
            for rows in sheet.iter_rows():
                reg_list.append(rows[0].value)
            
            if reg_number not in reg_list:
                sheet.cell(column=1, row=sheet.max_row+1, value=reg_number)
                sheet.cell(column=2, row=sheet.max_row, value=student_name)
                sheet.cell(column=3, row=sheet.max_row, value=student_class)
                sheet.cell(column=4, row=sheet.max_row, value=gender)
                sheet.cell(column=5, row=sheet.max_row, value=dob)
                sheet.cell(column=6, row=sheet.max_row, value=current_date)
                sheet.cell(column=7, row=sheet.max_row, value=student_religion)
                sheet.cell(column=8, row=sheet.max_row, value=student_skills)
                sheet.cell(column=9, row=sheet.max_row, value=fathers_name)
                sheet.cell(column=10, row=sheet.max_row, value=mothers_name)
                sheet.cell(column=11, row=sheet.max_row, value=fathers_occupation)
                sheet.cell(column=12, row=sheet.max_row, value=mothers_occupation)
                
                excel_file.save(excel_file_name)

                student_img_path = pathlib.Path('./student_images')
                if student_img_path.exists():
                    pass
                else:
                    os.mkdir('./student_images')
                try:
                    profile_image.save('./student_images/student_' + str(reg_number) + '_image' + '.png')
                except AttributeError:
                    pass

                c_file = pathlib.Path('./form_four_certificate')
                if c_file.exists():
                    pass
                else:
                    os.mkdir('./form_four_certificate')
                new_path = os.path.abspath('./form_four_certificate')
                shutil.copy(f_certificate, new_path)

                splited_old_path = f_certificate.split('/')
                file_name = splited_old_path[-1]
                file_extension = file_name.split('.')
                old_f_extension = file_extension[-1]
                old_path = f'{new_path}/{file_name}'
                new_f_path = f'{new_path}/student_{str(reg_number)}_form_four_certificate.jpg'
                os.rename(old_path, new_f_path)

                b_file = pathlib.Path('./birth_certificates')
                if b_file.exists():
                    pass
                else:
                    os.mkdir('./birth_certificates')
                new_b_path = os.path.abspath('./birth_certificates')
                shutil.copy(birth_certificate, new_b_path)

                splited_old_path = birth_certificate.split('/')
                file_name = splited_old_path[-1]
                file_extension = file_name.split('.')
                old_f_extension = file_extension[-1]
                old_path = f'{new_b_path}/{file_name}'
                new_f_path = f'{new_b_path}/student_{str(reg_number)}_birth_certificate.jpg'
                os.rename(old_path, new_f_path)

                messagebox.showinfo('Success', 'Data was saved Successful!')
                
            else:
                messagebox.showinfo('waring', 'The data your trying to save was already saved!')

    except:
        pass                 
        