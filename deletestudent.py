import openpyxl
import os
import pathlib
from tkinter import messagebox


def delete(registration_no, registration_no_entry):
    try:
        registration_no_entry.config(state='normal')
        reg_number = registration_no.get()
        empy_space = ' '

        excel_file_name = "./documents/Student_Registration_data.xlsx"
        excel_path = pathlib.Path('./documents')
        if excel_path.exists():
            pass
        else:
            os.mkdir('./documents')

        excel_file = openpyxl.load_workbook(excel_file_name)
        sheet = excel_file.active

        reg_list = []
        for rows in sheet.iter_rows():
            reg_list.append(rows[0].value)

        sheet.cell(column=1, row=reg_number + 1, value=empy_space)
        sheet.cell(column=2, row=reg_number + 1, value=empy_space)
        sheet.cell(column=3, row=reg_number + 1, value=empy_space)
        sheet.cell(column=4, row=reg_number + 1, value=empy_space)
        sheet.cell(column=5, row=reg_number + 1, value=empy_space)
        sheet.cell(column=6, row=reg_number + 1, value=empy_space)
        sheet.cell(column=7, row=reg_number + 1, value=empy_space)
        sheet.cell(column=8, row=reg_number + 1, value=empy_space)
        sheet.cell(column=9, row=reg_number + 1, value=empy_space)
        sheet.cell(column=10, row=reg_number + 1, value=empy_space)
        sheet.cell(column=11, row=reg_number + 1, value=empy_space)
        sheet.cell(column=12, row=reg_number + 1, value=empy_space)

        excel_file.save(excel_file_name)

        student_img_path = pathlib.Path('./student_images')
        if student_img_path.exists():
            pass
        else:
            os.mkdir('./student_images')
        try:
            os.remove('./student_images/student_' + str(reg_number) + '_image' + '.png')
        except AttributeError:
            pass
        messagebox.showinfo('Success', 'Data was deleted Successful!')

        c_file = pathlib.Path('./form_four_certificate')
        if c_file.exists():
            pass
        else:
            os.mkdir('./form_four_certificate')
        new_path = os.path.abspath('form_four_certificate')

        new_f_path = f'{new_path}/student_{str(reg_number)}_form_four_certificate.jpg'
        os.remove(new_f_path)

        b_file = pathlib.Path('./birth_certificates')
        if b_file.exists():
            pass
        else:
            os.mkdir('./birth_certificates')
        new_b_path = os.path.abspath('./birth_certificates')

        new_f_path = f'{new_b_path}/student_{str(reg_number)}_birth_certificate.jpg'
        os.remove(new_f_path)

        messagebox.showinfo('Success', 'Data was deleted Successful!')

    except:
        pass
