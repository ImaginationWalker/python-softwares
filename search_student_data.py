import pathlib
from tkinter import *
import openpyxl
from PIL import Image, ImageTk
from tkinter import messagebox


def search_data(search_value, registration_no, current_day_date, student, class_selection_combo, date_of_birth,
                religion, radio, skills, father_name, father_occupation, mother_name, mother_occupation, placeholder,
                from_four_certificate_entry, b_certificate_entry):
    try:
        excel_file = openpyxl.load_workbook('./documents/Student_Registration_data.xlsx')
        sheet = excel_file.active
        
        if search_value.get() >= 1:
            search_num = int(search_value.get()) + 1
            
            try:
                reg_number = sheet.cell(column=1, row=search_num).value
                student_name = sheet.cell(column=2, row=search_num).value
                student_class = sheet.cell(column=3, row=search_num).value
                student_gender = sheet.cell(column=4, row=search_num).value
                dob = sheet.cell(column=5, row=search_num).value
                current_day = sheet.cell(column=6, row=search_num).value
                student_religion = sheet.cell(column=7, row=search_num).value
                student_skills = sheet.cell(column=8, row=search_num).value
                fathers_name = sheet.cell(column=9, row=search_num).value
                mothers_name = sheet.cell(column=10, row=search_num).value
                fathers_occupation = sheet.cell(column=11, row=search_num).value
                mothers_occupation = sheet.cell(column=12, row=search_num).value

                registration_no.set(reg_number)
                student.set(student_name)
                class_selection_combo.set(student_class)
                date_of_birth.set(dob)
                current_day_date.set(current_day)
                religion.set(student_religion)
                skills.set(student_skills)
                father_name.set(fathers_name)
                mother_name.set(mothers_name)
                father_occupation.set(fathers_occupation)
                mother_occupation.set(mothers_occupation)
                file_path = f'./form_four_certificate/student_{str(reg_number)}_form_four_certificate.jpg'
                if pathlib.Path(file_path).exists():
                    from_four_certificate_entry.config(state='normal')
                    from_four_certificate_entry.insert(0, file_path)
                    from_four_certificate_entry.config(state='disabled')

                file_path1 = f'./birth_certificates/student_{str(reg_number)}_birth_certificate.jpg'
                if pathlib.Path(file_path1).exists():
                    b_certificate_entry.config(state='normal')
                    b_certificate_entry.insert(0, file_path1)
                    b_certificate_entry.config(state='disabled')

                if student_gender == "Male":
                    radio.set(1)
                if student_gender == 'Female':
                    radio.set(2)

                try:
                    profile_image = Image.open('./student_images/student_' + str(reg_number) + '_image' + '.png')
                    resized_image = profile_image.resize((248, 248))
                    student_image = ImageTk.PhotoImage(resized_image)
                    placeholder.config(image=student_image)
                    placeholder.image = student_image
                except FileNotFoundError:
                    messagebox.showinfo('No Image', f'No image for student with registration number {search_num-1}')
                                                                        
            except:
                messagebox.showinfo('No value', f'No info for Student with registration No: {search_num - 1}')
        
        else:
            messagebox.showinfo('No value', f'No info for Student with registration No: {search_value.get()}')

    except:
        pass                 
                      
    