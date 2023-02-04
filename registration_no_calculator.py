import openpyxl


def reg_no_calc():
    excel_file = openpyxl.load_workbook('./documents/Student_Registration_data.xlsx')
    sheet = excel_file.active
    row_value = sheet.max_row
    
    max_row_value = sheet.cell(row=row_value, column=1).value
   
    try:
        return max_row_value+1
      
    except TypeError:
        return 1     
